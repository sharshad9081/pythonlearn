# -*- coding: utf-8 -*-
"""
Created on Fri Nov 29 10:47:31 2019

@author: Rivatech
"""

from openpyxl import Workbook

wb = Workbook()

ws = wb.active

ws["A1"] = "test"
ws.append([1, 2, 3])

import datetime
ws["A2"] = datetime.datetime.now()

wb.save("sample.xlsx")