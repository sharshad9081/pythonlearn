# -*- coding: utf-8 -*-
"""
Created on Thu Nov 28 16:15:19 2019

@author: Rivatech
"""

import pymysql
from openpyxl import Workbook
import datetime
import time

"""
db = pymysql.connect(host = 'localhost', port = 3333, user = 'root', password = 'mysqlpassword')
print(db)
db.close()
print(db)



with pymysql.connect(host='localhost', port=3333, user='root', password='mysqlpassword', database="jpmc1") as db1:
#    print(db1)
#    query = "select * from emp"
#    db1.execute(query)
#    for record in db1.fetchall():
#        print(record)
    
#    query1 = "insert into emp values('{}', '{}')".format('tab', 'nel')
#    db1.execute(query1)
#    print(db1.rowcount, " inserted")
    
    query2 = "select * from emp"
    db1.execute(query2)
    for record in db1.fetchall():
        print(record)
        
        
"""

filename = time.strftime("%d_%b_%Y.xlsx")
connection = pymysql.connect(host='localhost', 
                                            port=3333, user='root', 
                                            password='mysqlpassword', 
                                            database="jpmc1")
try:
    wb = Workbook()
    ws = wb.active
#    ws["A1"] = ''
#    ws["A3"] = datetime.datetime.now()
#    with connection.cursor() as cursor:
#        sql = "INSERT INTO `emp` (`name`, `addr`) VALUES(%s, %s)"
#        cursor.execute(sql, ('par', 'dvg'))
#    connection.commit()
    
#    with connection.cursor() as cursor:
#        sql = "SELECT `name`, `addr` FROM `emp` where `addr`=%s"
#        cursor.execute(sql, ('rnr'))
#        result = cursor.fetchone()
#        print(result)
    
    with connection.cursor() as cursor:
        sql = "SELECT * FROM `emp`"
        cursor.execute(sql)
        result = cursor.fetchall()
        for record in result:
            ws.append(record)
#            print(record)
except:
    
finally:
#    wb.save("sample.xlsx")
    wb.save(filename)
    connection.close()





